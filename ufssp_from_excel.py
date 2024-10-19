import openpyxl as op

def take_settings(p_setting_path, p_setting_name):
	f = open(p_setting_path, 'r', encoding = "utf-8")
	s = ''
	for line in f:
		if line.find(p_setting_name) != -1:
			flag = False
			for i in range(len(line)):
				if line[i] == '\n':
					break

				if flag == True:
					s = s + line[i]

				if flag == False:
					if line[i] == '=':
						flag = True
	f.close()
	return s


def load_from_excel(p_setting_path, p_excel_path, p_vksp):
	wb = op.load_workbook(p_excel_path)
	ws = wb['Лист1']
	n = -1

	start_row = int(take_settings(p_setting_path, "начальная_строка"))
	start_col = int(take_settings(p_setting_path, "начальный_столбец"))

	for row in ws.iter_rows(min_row = start_row, min_col = start_col, max_col = start_col, max_row = 15000):
		if row[0].value == int(p_vksp):
			n = row[0].row
			break

	for row in ws.iter_rows(min_row = start_row, min_col = start_col, max_col = start_col, max_row = 15000):
		if type(row[0].value) == type(int()):
			if row[0].value == int(p_vksp):
				n = row[0].row
				print("type = int")
				break
		elif type(row[0].value) == type(str()):
			if len(row[0].value) <= 8: # костыль
				if row[0].value.find(str(p_vksp)) != -1:
					n = row[0].row
					print("type = str")
					break
		else:
			pass

	if n == -1:
		print("Неуспешная загрузка по ВКСП")
		return None, None

	print("Загрузка по ВКСП прошла успешно")
	d1 = ws.cell(row = n - 3, column = start_col).value
	d2 = ws.cell(row = n - 2, column = start_col).value
	return d1, d2
