from tkinter import *
import openpyxl as op
import algorithms
import ufssp_from_excel as ufe
import os

#PATH_UFSSP = "excel/tca.xlsx"
PATH_VALUTA = "excel/valuta.xlsx"
PATH_SETTING = "settings/settings.txt"


STANDART_FONT = ("Arial", 11)

MAS_LABEL_TEXT_FIRST_WINDOW = [ "Название УФССП",
								"Адрес УФССП",
								"Дата создания документа",
								"Дата принятия к исполнению",
								"Номер ИП",
								"ФИО ФЛ (ИП)",
								"Дата рождения ФЛ (ИП)",
								"Номер запроса"]

MAS_ENTRY_DEFAULT_TEXT_FOR_TEST_FIRST_WINDOW = ["TEST_УФССП по Омской области МОСП по ОВИП",
												"TEST_644053, г. Москва, ул. Пушкина, д. 85, корп. 3",
												"TEST_01.01.2024",
												"TEST_02.02.2024",
												"TEST_1337/228/1337-ИП",
												"TEST СЕМЕНОВ ИВАН АНДРЕЕВИЧ",
												"TEST_03.03.2024",
												"TEST_00000000-228b-a1bd-1234-019118b71b78"]
MAS_LABEL_TEXT_FIRST_WINDOW_UL_DIFFERENCE = ["Название ООО", "ИНН"]
MAS_ENTRY_DEFAULT_TEXT_FOR_TEST_FIRST_WINDOW_UL_DIFFERENCE = ["TEST_ООО СУПЕРХЛЕБУШЕК", "TEST_13372281337"]


TEXT_LSES_DEFAULT = """TEST_Номер счета: 40 На счете: 0
Номер счета: 411 На счете: 2
Номер счета: 4222 На счете: 0
Номер счета: 43333 На счете: 322
Номер счета: 444444 На счете: 0
Номер счета: 4555555 На счете: -100.21
Номер счета: null На счете: null
Номер счета: 40802810200001529859 На счете: 0
Номер счета: null На счете: null"""

TEXT_LSES_DEFAULT = """TEST_Номер счета: 40802810200001529859 На счете: 0
Номер счета: 40802810200001529859 На счете: -100
Номер счета: 40802810200001529859 На счете: 0
Номер счета: 40802810200001529859 На счете: 0
Номер счета: 40802810200001529859 На счете: 0
Номер счета: 40802810200001529859 На счете: -100.21
Номер счета: null На счете: null
Номер счета: 40802810200001529859 На счете: 0
Номер счета: null На счете: null"""


SECOND_WINDOW_CREATE_DOC_STATUS_CONSTANT_ERROR = "Ошибка в создании документа"
SECOND_WINDOW_CREATE_DOC_STATUS_CONSTANT_COMPLETE = "Документ создан успешно"

class Program():
	obrabotka_type = None
	current_window = None

	win_first = None
	win_second = None

	lbl_mas = None
	entr_mas = None

	text_lses = None
	lbl_complete = None

	entr_second_mas = None
	second_window_create_doc_status = None

	lod = None

	def create_start_lod_and_text_lses(self):
		self.lod = []
		for i in range(len(MAS_ENTRY_DEFAULT_TEXT_FOR_TEST_FIRST_WINDOW)):
			if i == 5 or i == 6:
				if self.obrabotka_type == "FL_and_IP":
					s = MAS_ENTRY_DEFAULT_TEXT_FOR_TEST_FIRST_WINDOW[i]
				elif self.obrabotka_type == "UL":
					if i ==  5:
						s = MAS_ENTRY_DEFAULT_TEXT_FOR_TEST_FIRST_WINDOW_UL_DIFFERENCE[0]
					else:
						s = MAS_ENTRY_DEFAULT_TEXT_FOR_TEST_FIRST_WINDOW_UL_DIFFERENCE[1]
				else:
					s = "ERROR"
			else:
				s = MAS_ENTRY_DEFAULT_TEXT_FOR_TEST_FIRST_WINDOW[i]

			self.lod.append(s)
		self.lod.append(TEXT_LSES_DEFAULT)

		self.text_lses = TEXT_LSES_DEFAULT



	def __init__(self):
		self.obrabotka_type = "FL_and_IP"

		self.create_start_lod_and_text_lses()
		self.create_first_window()
		
	# Функция для создания текста с определенными настройками позиции
	def create_label_for_first_window(self, p_i, p_text):
		lbl = Label(text = p_text, font = STANDART_FONT)
		lbl.pack(anchor = NW, padx = [5, 20])
		return lbl

	# Функция для создания поля для ввода с определенными настройками позиции
	def create_entry_for_first_window(self, p_i, p_text):
		entr = Entry(font = STANDART_FONT)
		entr.pack(anchor = NW, fill = X, padx = [5, 20], pady = [3, 13], ipadx = 2, ipady = 1)
		entr.insert(0, p_text)
		return entr

	def create_document(self):
		ls = []
		summa = []
		value_type = []
		for i in range(len(self.entr_second_mas)):
			ls.append(self.entr_second_mas[i][0].get())
			summa.append(self.entr_second_mas[i][1].get())
			value_type.append(self.entr_second_mas[i][2].get())
		self.second_window_create_doc_status.config(text = SECOND_WINDOW_CREATE_DOC_STATUS_CONSTANT_ERROR)
		algorithms.all_together(self.obrabotka_type, self.lod, ls, summa, value_type)
		self.second_window_create_doc_status.config(text = SECOND_WINDOW_CREATE_DOC_STATUS_CONSTANT_COMPLETE)
	
	def take_currency_from_nomer_ls(self, ls):
		wb = op.load_workbook(PATH_VALUTA)
		ws = wb['Лист1']
		n = -1
		for row in ws.iter_rows(min_row = 2, max_row = 254, min_col = 4, max_col = 4):
			s = str(row[0].value)
			if len(s) == 2:
				s = '0' + s
			elif len(s) == 1:
				s = '00' + s

			if s == ls[5:8]:
				n = row[0].row
				break
		if n == -1:
			return "ERROR"
		
		d = ws.cell(row = n, column = 3).value
		return d


	def from_excel(self, p_lbl_status, p_vksp):
		excel_path = ufe.take_settings(PATH_SETTING, "расположение_файла_с_уфссп")
		d1, d2 = ufe.load_from_excel(PATH_SETTING, excel_path, p_vksp)
		
		if d1 == None:
			p_lbl_status.config(text = "Неуспешная загрузка по ВКСП")
			return

		p_lbl_status.config(text = "Загрузка по ВКСП прошла успешно")
		self.entr_mas[0].delete(0, last = END)
		self.entr_mas[1].delete(0, last = END)
		self.entr_mas[0].insert(0, d1)
		self.entr_mas[1].insert(0, d2)

	def change_to_FL_and_IP(self):
		if self.obrabotka_type == "FL_and_IP":
			return

		self.obrabotka_type = "FL_and_IP"
		self.current_window.destroy()

		self.create_start_lod_and_text_lses()
		self.create_first_window()



	def change_to_UL(self):
		if self.obrabotka_type == "UL":
			return

		self.obrabotka_type = "UL"
		self.current_window.destroy()

		self.create_start_lod_and_text_lses()
		self.create_first_window()


	def create_first_window(self):
		self.win_first = Tk()
		self.current_window = self.win_first
		
		if self.obrabotka_type == "FL_and_IP":
			self.win_first.title("Главное окно обработки ФЛ и ИП")
		elif self.obrabotka_type == "UL":
			self.win_first.title("Главное окно обработки ЮЛ")
		else:
			self.win_first.title("ERROR")

		main_menu = Menu()
		file_menu = Menu(tearoff = False)


		lbl_str_fl_and_ip = "Арест на ФЛ или ИП"
		lbl_str_ul = "Арест на ЮЛ"

		if self.obrabotka_type == "FL_and_IP":
			lbl_str_fl_and_ip = lbl_str_fl_and_ip + "*"
		elif self.obrabotka_type == "UL":
			lbl_str_ul = lbl_str_ul + "*"
		else:
			lbl_str_fl_and_ip = "ERROR"
			lbl_str_ul = "ERROR"

		file_menu.add_command(label = lbl_str_fl_and_ip, command = lambda: self.change_to_FL_and_IP())
		file_menu.add_command(label = lbl_str_ul, command = lambda: self.change_to_UL())
		main_menu.add_cascade(label = "Сменить арест", menu = file_menu)
		self.win_first.config(menu = main_menu)



		self.lbl_mas = []
		self.entr_mas = []
		for i in range(len(MAS_LABEL_TEXT_FIRST_WINDOW)):
			if i == 2:
				
				lbl_status = Label(text = "< Статус загрузки по ВКСП >", font = STANDART_FONT)
				lbl_status.pack(anchor = NE, padx = [5, 20])

				entry_code = Entry(font = STANDART_FONT)
				entry_code.pack(anchor = NE, padx = [5, 20], pady = [3, 13], ipadx = 2, ipady = 1)
				entry_code.insert(0, "000000")

				btn = Button(text = "Загрузить по ВКСП из таблицы", command = lambda: self.from_excel(lbl_status, entry_code.get()))
				btn.pack(anchor = NE, padx = [5, 20])

			if i == 5 or i == 6:
				if self.obrabotka_type == "FL_and_IP":
					lbl_take_str = MAS_LABEL_TEXT_FIRST_WINDOW[i]
				elif self.obrabotka_type == "UL":
					if i == 5:
						lbl_take_str = MAS_LABEL_TEXT_FIRST_WINDOW_UL_DIFFERENCE[0]
					else:
						lbl_take_str = MAS_LABEL_TEXT_FIRST_WINDOW_UL_DIFFERENCE[1]
				else:
					lbl_take_str = "ERROR"
			else:
				lbl_take_str = MAS_LABEL_TEXT_FIRST_WINDOW[i]

			lbl = self.create_label_for_first_window(i + i*2, lbl_take_str)
			self.lbl_mas.append(lbl)

			entr = self.create_entry_for_first_window(i + 1+i*2, self.lod[i])
			self.entr_mas.append(entr)
			

		self.entr_mas[0].config(width = 100)

		self.text_lses = Text(height = 10, width = 60, font = STANDART_FONT)
		self.text_lses.pack(anchor = NW)
		self.text_lses.insert("1.0", self.lod[8])

		self.lbl_complete = Label(text = "")
		self.lbl_complete.pack(padx = 20)

		btn = Button(text = "Загрузить данные", command = lambda: self.check_for_perehod_to_second_window())
		btn.pack(anchor = S, pady = 10, padx = [5, 20])

		self.win_first.mainloop()
		

	def check_for_perehod_to_second_window(self):
		flag = True
		for i in range(len(self.entr_mas)):
			self.lod[i] = self.entr_mas[i].get()
			if self.lod[i] == '' or self.lod[i] == MAS_ENTRY_DEFAULT_TEXT_FOR_TEST_FIRST_WINDOW[i]:
				flag = False

		self.lod[8] = self.text_lses.get("1.0", "end")
		if self.lod[8] == TEXT_LSES_DEFAULT:
			flag = False

		flag = True
		if flag == False:
			self.lbl_complete.config(text = 'Не все данные были введены')
			return
		else:
			self.lbl_complete.config(text = 'Всё успешно')
			self.win_first.destroy()
			self.create_second_window()


	def back_to_first_window(self):
		self.win_second.destroy()
		self.create_first_window()

	def open_output_docx(self):
		if self.second_window_create_doc_status["text"] != SECOND_WINDOW_CREATE_DOC_STATUS_CONSTANT_COMPLETE:
			return

		if self.obrabotka_type == "FL_and_IP":
			s = algorithms.path_answer_FL_and_IP.replace('/', '\\')
			os.startfile(s)
		elif self.obrabotka_type == "UL":
			s = algorithms.path_answer_UL.replace('/', '\\')
			os.startfile(s)
		else:
			print("ERROR. Неверный obrabotka_type")



	def create_second_window(self):
		self.win_second = Tk()
		self.current_window = self.win_second

		if self.obrabotka_type == "FL_and_IP":
			self.win_second.title("Окно обработки номеров ЛС ФЛ или ИП")
		elif self.obrabotka_type == "UL":
			self.win_second.title("Окно обработки номеров ЛС ЮЛ")
		else:
			self.win_second.title("ERROR")

		btn = Button(text = "Вернуться назад", command = lambda: self.back_to_first_window())
		btn.grid(row = 0, column = 0, sticky = NW, padx = 10, pady = 5) 

		ls, summa = algorithms.take_lists_ls_and_summa(self.lod[8])

		self.entr_second_mas = []
		for r in range(len(ls)):
			self.entr_second_mas.append([])

			entr = Entry(font = STANDART_FONT, width = 25)
			entr.grid(row = r+1, column = 0,  sticky = NW, padx = [10, 20], pady = 10, ipadx = 6, ipady = 3)
			entr.insert(0, ls[r])
			self.entr_second_mas[r].append(entr)

			entr = Entry(font = STANDART_FONT, width = 10)
			entr.grid(row = r+1, column = 1, sticky = EW, pady = 10, ipadx = 6, ipady = 3)
			entr.insert(0, summa[r])
			self.entr_second_mas[r].append(entr)
			
			entr = Entry(font = STANDART_FONT, width = 7)
			entr.grid(row = r+1, column = 2, sticky = NE, padx = [20, 10], pady = 10, ipadx = 6, ipady = 3)
			entr.insert(0, self.take_currency_from_nomer_ls(ls[r]))
			self.entr_second_mas[r].append(entr)
	
		self.second_window_create_doc_status = Label(text = "", font = STANDART_FONT)
		self.second_window_create_doc_status.grid(row = len(ls)+1, column = 1, sticky = NE, padx = 10, pady = 10)

		btn2 = Button(text = "Создать документ", command = lambda: self.create_document())
		btn2.grid(row = len(ls)+1, column = 0, sticky = NW, padx = 10, pady = 10) 

		btn_open_output_docx = Button(text = "Открыть полученный файл", command = lambda: self.open_output_docx())
		btn_open_output_docx.grid(row = len(ls)+1, column = 2, sticky = NE, padx = 10, pady = 10) 


		self.win_second.mainloop()


		

p1 = Program()